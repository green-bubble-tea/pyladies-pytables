from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.cell import get_column_letter, column_index_from_string, coordinate_from_string
from pyexcelerate import Workbook
from datetime import datetime

import codecs
import csv
import os

light_yellow = PatternFill(fgColor="FFF2CC", fill_type="solid")
gold_yellow = PatternFill(fgColor="FFD966", fill_type="solid")
border = Side(border_style="thick", color='000000')
border_thin = Side(border_style="thin", color='000000')


def draw_border(start, end, sheet):
    col_start, ind_start = coordinate_from_string(start)
    col_end, ind_end = coordinate_from_string(end)

    left = [f"{col_start}{i}" for i in range(ind_start, ind_end + 1)]
    top = [f"{get_column_letter(i)}{ind_start}" for i in
           range(column_index_from_string(col_start), column_index_from_string(col_end) + 1)]
    bottom = [f"{get_column_letter(i)}{ind_end}" for i in
              range(column_index_from_string(col_start), column_index_from_string(col_end) + 1)]
    right = [f"{col_end}{i}" for i in range(ind_start, ind_end + 1)]

    top_left = f"{col_start}{ind_start}"
    sheet[top_left].border = Border(top=border, left=border)
    bottom_left = f"{col_start}{ind_end}"
    sheet[bottom_left].border = Border(bottom=border, left=border)
    top_right = f"{col_end}{ind_start}"
    sheet[top_right].border = Border(top=border, right=border)
    bottom_right = f"{col_end}{ind_end}"
    sheet[bottom_right].border = Border(bottom=border, right=border)

    if top_left == bottom_left:
        sheet[top_left].border = Border(top=border, bottom=border, left=border)
    if top_right == bottom_right:
        sheet[top_right].border = Border(top=border, bottom=border, right=border)

    for i in top:
        if i not in [top_left, top_right]:
            sheet[i].border = Border(top=border)
    for i in bottom:
        if i not in [bottom_left, bottom_right]:
            sheet[i].border = Border(bottom=border)
    for i in left:
        if i not in [top_left, bottom_left]:
            sheet[i].border = Border(left=border)
    for i in right:
        if i not in [top_right, bottom_right]:
            sheet[i].border = Border(right=border)


def exercise_2():
    dir_name = "lesson_02"
    file_name = "table_"
    with codecs.open(os.path.join(dir_name, "seznam_subjektu.csv"), "r", encoding="ISO 8859-2") as sub_file:
        sub_list = csv.DictReader(sub_file, delimiter=";")
        for i in sub_list:
            ico = i['ICO']
            fname = os.path.join(dir_name, f"{file_name}{ico}.xlsx")
            header = "Karta danoveho subjektu"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = i['ICO']
            h = sheet["A1"]
            h.value = header
            h.font = Font(name="Calibri", size=12, bold=True)
            sheet.merge_cells("A1:E1")
            h.alignment = Alignment(horizontal="center", vertical="center")
            sheet.row_dimensions[1].height = 42
            for letter in "ABCDE":
                if letter == "C":
                    sheet.column_dimensions[letter].width = 5
                else:
                    sheet.column_dimensions[letter].width = 25
            sheet["A2"] = "Vyhotvila"
            sheet["B2"] = "XYZ"
            sheet["D2"] = "Vytvoreno"
            sheet["E2"] = datetime.today().strftime('%d-%m-%Y')
            sheet.row_dimensions[2].height = 30
            for cell in sheet["2:2"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Calibri", size=12, bold=True)
            sheet["A4"] = "ICO"
            sheet["B4"] = i['ICO']
            sheet["D4"] = "Nazev firmy"
            sheet["E4"] = i["Nazev firmy"]
            for row in range(4, 12):
                sheet.row_dimensions[row].height = 30
            sheet["A6"] = "Ulice"
            sheet["B6"] = i["ulice"]
            sheet["D6"] = "Jednatel"
            sheet["E6"] = i["jednatel firmy"]

            sheet["A7"] = "Mesto"
            sheet["B7"] = i["Mesto"]
            sheet["D7"] = "Pocet zamestnancu"
            sheet["E7"] = i["pocet zamestnancu"]

            sheet["A8"] = "PSC"
            sheet["B8"] = i["PSC"]
            sheet["D8"] = "Zakladni kapital"
            sheet["E8"] = i["zakladni kapital"]

            sheet["A10"] = "Platce DPH"
            sheet["B10"] = i["platce DPH"]
            sheet["D10"] = "DIC"
            sheet["E10"] = i["DIC"]

            sheet["D11"] = "Registrace OD"
            sheet["E11"] = i["Registrace od"]
            sheet["D11"].fill = light_yellow
            sheet["E11"].fill = light_yellow

            for index, row in enumerate(sheet["A6:E10"]):
                for cell in row:
                    if index in [0, 2] and cell.value is not None and cell.column_letter != "C":
                        cell.fill = light_yellow

            # range A6:B8
            draw_border("A6", "B8", sheet)
            draw_border("D6", "E8", sheet)
            draw_border("A10", "B10", sheet)
            draw_border("D10", "E11", sheet)

            for cell in sheet["4:4"]:
                cell.font = Font(name="Calibri", size=12, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.column_letter != "C":
                    cell.border = Border(left=border,
                                         right=border,
                                         top=border,
                                         bottom=border)
                if cell.column_letter in "AD":
                    cell.fill = gold_yellow
                elif cell.column_letter in "BE":
                    cell.fill = light_yellow

            for i in range(1, 101):
                cell = sheet[f"{get_column_letter(i)}13"]
                if i % 2 != 0:
                    cell.fill = light_yellow

            workbook.save(fname)


if __name__ == "__main__":
    exercise_2()
