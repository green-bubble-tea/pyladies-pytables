# pyt-able.py
# module for EXCEL exports

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


def create_demo(filename="example.xlsx", title="Example"):
    workbook = Workbook()

    try:
        worksheet = workbook.active
        worksheet.title = title

        # merge cells for header
        worksheet.merge_cells('A1:F1')
        worksheet["A1"] = title

        set_columns_width(worksheet, 25, for_all=True, col_max=6)
        set_rows_height(worksheet, 25, for_all=True, row_max=30)

        set_area_alignment(worksheet, 'A1:A30', horizontal='right', indent=1)

    finally:
        workbook.save(filename=filename)


def set_columns_width(worksheet, widths, for_all=False, col_min=1, col_max=100):
    if col_min > col_max or col_max < 1 or col_min < 1 or col_max > 100 or col_min > 100:
        raise AssertionError("Wrong col_min/col_max parameters")

    if type(widths) not in (list, tuple, int):
        if for_all:
            raise AssertionError("widths argument is wrong with for_all (must be list or tuple or int)")
        else:
            raise AssertionError("widths argument is wrong without for_all (must be list or tuple)")

    if for_all:
        if type(widths) in (list, tuple):
            for no in range(col_min, col_max+1):
                worksheet.column_dimensions[get_column_letter(no)].width = widths[no % len(widths)-1]

        elif type(widths) == int:
            for no in range(col_min, col_max + 1):
                worksheet.column_dimensions[get_column_letter(no)].width = widths

    else:
        if type(widths) in (list, tuple):
            for no, width in enumerate(widths, 1):
                worksheet.column_dimensions[get_column_letter(no)].width = width


def set_rows_height(worksheet, heights, for_all=False, row_min=1, row_max=100):
    if row_min > row_max or row_max < 1 or row_min < 1 or row_max > 100 or row_min > 100:
        raise AssertionError("Wrong row_min/row_max parameters")

    if type(heights) not in (list, tuple, int):
        if for_all:
            raise AssertionError("heights argument is wrong with for_all (must be list or tuple or int)")
        else:
            raise AssertionError("heights argument is wrong without for_all (must be list or tuple)")

    if for_all:
        if type(heights) in (list, tuple):
            for no in range(row_min, row_max+1):
                worksheet.row_dimensions[no].height = heights[no % len(heights)-1]

        elif type(heights) == int:
            for no in range(row_min, row_max + 1):
                worksheet.row_dimensions[no].height = heights

    else:
        if type(heights) in (list, tuple):
            for no, height in enumerate(heights, 1):
                worksheet.row_dimensions[no].height = height


def set_area_alignment(worksheet, area, horizontal='left', vertical='center', indent=0):
    for row in worksheet[f'{area}']:
        for cell in row:
            cell.alignment = Alignment(horizontal=horizontal, vertical=vertical, indent=indent)


def set_area_font(worksheet, area, size=12, font_type="Calibri", bold=False, italic=False, underscore=False):
    "set font to selected area"
    ...


create_demo()
